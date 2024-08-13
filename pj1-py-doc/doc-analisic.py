import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook

def select_file():
    root = tk.Tk()
    root.withdraw()  # Hide the root window

    file_path = filedialog.askopenfilename(
        title="Select an Excel File",
        filetypes=[("Excel Files", "*.xlsx")] #pre select type of files to show
    )

    return file_path

def get_cell_value(file_path, sheet_name, cell):
    # Load the workbook
    workbook = load_workbook(filename=file_path, data_only=True)
    
    # Check if the sheet exists
    if sheet_name not in workbook.sheetnames:
        print(f"Sheet '{sheet_name}' not found in the document.")
        return None
    
    sheet = workbook[sheet_name]
    
    # Get the value from the specified cell
    cell_value = sheet[cell].value
    return cell_value
if __name__ == "__main__": #???????
    file_path = select_file()
ls =[]
inpt = input("'pass' to stop enter to continue: ")
while inpt != "pass":
    inpt = input("choose a cell: ")
    if inpt != "pass":
        ls.append(inpt)
        
for cell in ls:
    if file_path:
        # Specify the sheet name and cell to read
        sheet_name = 'Sheet1'  # Adjust to the sheet name in your file
        value = get_cell_value(file_path, sheet_name, cell)
        if value is not None:
            print(f"Value in cell {cell}: {value}")
    else:
        print("No file selected.")


#in terms of cell selection, it works for lower case letters also
#because you can have more that one sheet in a document, you need to spesify wich sheet you want infoirmatin from 